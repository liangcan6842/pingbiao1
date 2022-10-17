# encoding=utf-8
# @Time    :2022/10/10 9:52
# @Author  :梁灿
# @File    :Package_file_method.py
#封装公用方法：读取yml文件、excel文件
import os

import pytest
import yaml
import openpyxl
from openpyxl import load_workbook

path = 'C:\\Users\\Administrator\\Desktop\\pingbiao-API\\data'
# 读取data文件目录下的yaml测试数据，需要传入yaml文件的名字，调用方法，并返回
def read_yaml(yaml_name):
    with open(path + '\\' + yaml_name,encoding='utf-8') as yml:
        yml_data = yaml.load(yml,yaml.FullLoader)
        print(yml_data)
    return yml_data

# def read_excel():
#     wb = openpyxl.load_workbook("C:/Users/Admin/Desktop/test.xlsx")
#     wb.active
#     wb.sheetnames
#     print("excel文件的所有sheet名字：%s" % wb.sheetnames)
#     ws = wb['Sheet1']
#     # 遍历接收到的内容
#     print("打印出有多少行：%s" % len(tuple(ws.rows)))
#     test_date = []
#     for x in range(2, len(tuple(ws.rows)) + 1):
#         testcase_data = []
#         for y in range(2, 4):
#             testcase_data.append(ws.cell(row=x, column=y).value)
#             print("打印每行的值：%s" % ws.cell(row=x, column=y).value)
#         test_date.append(testcase_data)
#     print("打印出test_date的值：%s" % test_date)
#     return test_date